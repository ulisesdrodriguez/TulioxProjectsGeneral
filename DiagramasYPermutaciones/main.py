import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from PIL import Image, ImageTk
from itertools import permutations



excel_filename = "datos.xlsx"


if os.path.exists(excel_filename):
    workbook = load_workbook(excel_filename)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Diagrama_Ger", "Diagrama_Ra_1", "Diagrama_Ra_2"])  # Agregar encabezados

# Número de fila actual
fila_actual = 2  # Empezar después de los encabezados

# Función para guardar datos en Excel
def guardar_datos(entries, column):
    global fila_actual
    valores = [entry.get() for entry in entries]
    concatenated_values = ' '.join(valores)

    if concatenated_values.strip():  # Asegurarse de que no esté vacío
        sheet.cell(row=fila_actual, column=column, value=concatenated_values)
        if column == 3:  # Si estamos en la última columna, avanzamos a la siguiente fila
            fila_actual += 1
        workbook.save(excel_filename)
        messagebox.showinfo("Guardado", "Datos guardados correctamente")
        for entry in entries:
            entry.delete(0, tk.END)
    else:
        messagebox.showwarning("Advertencia", "Por favor, completa al menos un campo")

# Función para crear una ventana con el formulario
def crear_ventana(coords, titulo, column):
    ventana = tk.Toplevel()
    ventana.title(titulo)
    ventana.configure(bg="#117864")  # Cambiar el color de fondo de la ventana secundaria al color RGB especificado

    entries = []
    for r, c in coords:
        entry = tk.Entry(ventana, width=3, bg="gray20", fg="white")  # Cambiar el fondo de las entradas a gris oscuro y el texto a blanco
        entry.grid(row=r, column=c, padx=5, pady=5)
        entries.append(entry)

    boton_guardar = tk.Button(ventana, text="Guardar", command=lambda: guardar_datos(entries, column), bg="gray30", fg="white")  # Cambiar el fondo del botón a gris oscuro y el texto a blanco
    boton_guardar.grid(row=14, column=6, columnspan=3, pady=10)


def mostrar_permutaciones():
    # Leer datos de las columnas del archivo Excel
    diagrama_ger = sheet["A"][1:]  # Columna Diagrama_Ger
    diagrama_ra_1 = sheet["B"][1:]  # Columna Diagrama_Ra_1
    diagrama_ra_2 = sheet["C"][1:]  # Columna Diagrama_Ra_2

    # Generar permutaciones de los registros
    permutaciones_ger = {permutacion for celda in diagrama_ger for permutacion in permutations(celda.value)}
    permutaciones_ra_1 = {permutacion for celda in diagrama_ra_1 for permutacion in permutations(celda.value)}
    permutaciones_ra_2 = {permutacion for celda in diagrama_ra_2 for permutacion in permutations(celda.value)}

    # Encontrar permutaciones comunes entre las tres columnas
    permutaciones_comunes = permutaciones_ger & permutaciones_ra_1 & permutaciones_ra_2

    # Crear una ventana para mostrar los resultados
    ventana_resultados = tk.Toplevel()
    ventana_resultados.title("Permutaciones Comunes")
    ventana_resultados.configure(bg="#ebf5fb")

    # Mostrar las permutaciones comunes en la ventana
    if permutaciones_comunes:
        resultados_texto = tk.Text(ventana_resultados, width=50, height=20, bg="gray20", fg="white")
        resultados_texto.pack(pady=20)
        for permutacion in permutaciones_comunes:
            resultados_texto.insert(tk.END, ''.join(permutacion) + '\n')
    else:
        tk.Label(ventana_resultados, text="No se encontraron permutaciones comunes.", bg="#ebf5fb", fg="black").pack(pady=20)


# Coordenadas del primer formulario en forma de rombo
coords_1 = [
    (1, 6),
    (2, 5), (2, 7),
    (3, 4), (3, 6), (3, 8),
    (4, 3), (4, 5), (4, 7), (4, 9),
    (5, 2), (5, 4), (5, 6), (5, 8), (5, 10),
    (6, 1), (6, 3), (6, 5), (6, 7), (6, 9), (6, 11),
    (7, 0), (7, 2), (7, 4), (7, 6), (7, 8), (7, 10), (7, 12),
    (8, 1), (8, 3), (8, 9), (8, 11),(9, 2), (9, 4),
    (9, 6), (9, 8), (9, 10), (10, 3), (10, 5), (10, 7),
    (10, 9), (11, 4), (11, 6),(11, 8), (12, 5),
    (12, 7), (13, 6)
]

# Coordenadas del segundo formulario (modificadas)
coords_2 = [
    (0, 3), (0, 6), (0, 9),
    (2, 6),
    (3, 3), (3, 4), (3, 8),(3, 9),
    (4, 3), (4, 9),
    (6, 6),
    (7, 5), (7, 7),
    (8, 4), (8, 6), (8, 8),
    (9, 3), (9, 5), (9, 7), (9, 9),
    (10, 2), (10, 4), (10, 6), (10, 8), (10, 10),
    (11, 1), (11, 3), (11, 5), (11, 7), (11, 9), (11, 11)
]

# Coordenadas del tercer formulario (modificadas)
coords_3 = [
    (0, 3), (0, 6), (0, 9),
    (2, 6),
    (3, 3), (3, 4), (3, 8),(3, 9),
    (4, 3), (4, 9),
    (6, 6),
    (7, 5), (7, 7),
    (8, 4), (8, 6), (8, 8),
    (9, 3), (9, 5), (9, 7), (9, 9),
    (10, 2), (10, 4), (10, 6), (10, 8), (10, 10),
    (11, 1), (11, 3), (11, 5), (11, 7), (11, 9), (11, 11)
]

# Crear la ventana principal
ventana_principal = tk.Tk()
ventana_principal.title("App de Tuliox que nos dara money")
ventana_principal.geometry("600x750")  # Cambia el tamaño según tus necesidades
ventana_principal.configure(bg="#ebf5fb")  # Cambiar el color de fondo de la ventana principal al color RGB especificado

image_path = "excaliburicon.jpg" # Cambia esto a la ruta de tu imagen
image = Image.open(image_path)
photo = ImageTk.PhotoImage(image) # Crear un widget Label para la imagen
image_label = tk.Label(ventana_principal, image=photo)
image_label.place(relx=0.5, rely=0.5, anchor="center")

# Botones para abrir las otras ventanas
boton1 = tk.Button(ventana_principal, text="Diagrama Getse", command=lambda: crear_ventana(coords_1, "Diagrama Getse", 1))
boton1.pack(pady=10)

boton2 = tk.Button(ventana_principal, text="Diagrama Ra 1", command=lambda: crear_ventana(coords_2, "Diagrama Ra 1", 2))
boton2.pack(pady=10)

boton3 = tk.Button(ventana_principal, text="Diagrama Ra 2", command=lambda: crear_ventana(coords_3, "Diagrama Ra 2 ", 3))
boton3.pack(pady=10)

boton4 = tk.Button(ventana_principal, text="Permutaciones", command=mostrar_permutaciones)
boton4.pack(pady=10)
# Iniciar el bucle de la interfaz gráfica
ventana_principal.mainloop()
